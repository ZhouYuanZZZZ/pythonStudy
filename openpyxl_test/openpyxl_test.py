from openpyxl import Workbook
from openpyxl import load_workbook


def test0():
    wb = load_workbook(r'C:\Users\zy127\Desktop\test.xlsx')
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    for row in ws.iter_rows(min_row=1, max_row=13, min_col=1, max_col=4, values_only=True):
        for cell in row:
            print(type(cell))


test0()
